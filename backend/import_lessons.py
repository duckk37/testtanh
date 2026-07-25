import os
import re
import models
import openpyxl

def extract_youtube_id(url):
    if not url:
        return "dQw4w9WgXcQ"
    # Matches v=XXXX, or youtu.be/XXXX
    match = re.search(r'(?:v=|youtu\.be/)([^&]+)', url)
    if match:
        return match.group(1)
    return "dQw4w9WgXcQ"

def import_data(db):
    try:
        # Create the course if not exists
        course_title = "48 NGÀY LẤY GỐC TIẾNG ANH"
        course = db.query(models.Course).filter(models.Course.title == course_title).first()
        if not course:
            course = models.Course(
                title=course_title,
                description="Khóa học lấy lại nền tảng tiếng Anh trong 48 ngày. (Data imported from Google Sheet)",
                thumbnail="https://tienganhcomaiphuong.vn/wp-content/uploads/2021/08/pro3m.jpg",
                price=0.0
            )
            db.add(course)
            db.commit()
            db.refresh(course)
            
        print(f"Course ID: {course.id}")
        
        # Read the Excel file
        excel_path = os.path.join(os.path.dirname(__file__), 'data.xlsx')
        if not os.path.exists(excel_path):
            print("data.xlsx not found!")
            return

        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        start_row = 1
        for row in range(1, 10):
            val = sheet.cell(row=row, column=2).value
            if val and 'TÊN BÀI HỌC' in str(val):
                start_row = row + 1
                break
                
        print(f"Starting import from row {start_row}")
        
        count = 0
        updated = 0
        for row in range(start_row, sheet.max_row + 1):
            stt_val = sheet.cell(row=row, column=1).value
            title_val = sheet.cell(row=row, column=2).value
            
            if not stt_val or not title_val:
                continue
                
            try:
                order_index = int(stt_val)
            except ValueError:
                continue
                
            title = str(title_val).strip()
            
            # Extract hyperlink
            cell_c = sheet.cell(row=row, column=3)
            hyperlink = cell_c.hyperlink.target if cell_c.hyperlink else None
            youtube_id = extract_youtube_id(hyperlink)
            
            # Check if lesson already exists
            existing_lesson = db.query(models.Lesson).filter(
                models.Lesson.course_id == course.id,
                models.Lesson.order_index == order_index
            ).first()
            
            if not existing_lesson:
                lesson = models.Lesson(
                    course_id=course.id,
                    title=title,
                    youtube_id=youtube_id,
                    order_index=order_index,
                    passing_score_required=80
                )
                db.add(lesson)
                count += 1
            else:
                # Update existing lesson
                if existing_lesson.youtube_id != youtube_id and youtube_id != "dQw4w9WgXcQ":
                    existing_lesson.youtube_id = youtube_id
                    updated += 1
        
        db.commit()
        print(f"Successfully imported {count} new lessons, updated {updated} lessons.")
        
    except Exception as e:
        print(f"Error importing lessons: {e}")
