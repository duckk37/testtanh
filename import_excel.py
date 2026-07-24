import sys
import os
import re

try:
    import openpyxl
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    sys.exit(1)

# Add the backend directory to sys.path so we can import models and database
sys.path.append(os.path.join(os.path.dirname(__file__), 'backend'))

from database import SessionLocal, engine
import models

def extract_youtube_id(url):
    if not url:
        return None
    
    # https://www.youtube.com/watch?v=1b3mD3J0q3E
    # https://youtu.be/1b3mD3J0q3E
    youtube_regex = (
        r'(https?://)?(www\.)?'
        '(youtube|youtu|youtube-nocookie)\.(com|be)/'
        '(watch\?v=|embed/|v/|.+\?v=)?([^&=%\?]{11})'
    )
    
    match = re.search(youtube_regex, url)
    if match:
        return match.group(6)
    return None

def import_excel_data():
    wb = openpyxl.load_workbook("../48 NGÀY LẤY GỐC TIẾNG ANH.xlsx")
    sheet = wb.active

    # Find where the data starts
    start_row = 1
    for row in range(1, 20):
        cell_val = str(sheet.cell(row=row, column=1).value).strip()
        if cell_val == "STT":
            start_row = row + 1
            break
            
    print(f"Starting to process from row {start_row}")

    db = SessionLocal()
    try:
        course_title = "48 NGÀY LẤY GỐC TIẾNG ANH"
        course = db.query(models.Course).filter(models.Course.title == course_title).first()
        
        if not course:
            print("Course not found in database!")
            return

        count = 0
        for row in range(start_row, sheet.max_row + 1):
            stt_cell = sheet.cell(row=row, column=1).value
            if not stt_cell:
                continue
                
            try:
                order_index = int(str(stt_cell).strip())
            except ValueError:
                continue
                
            # Bài giảng is usually in column 3 (C)
            link_cell = sheet.cell(row=row, column=3)
            
            url = None
            if link_cell.hyperlink:
                url = link_cell.hyperlink.target
            elif str(link_cell.value).startswith("http"):
                url = link_cell.value
                
            if url:
                yt_id = extract_youtube_id(url)
                if yt_id:
                    # Update lesson
                    lesson = db.query(models.Lesson).filter(
                        models.Lesson.course_id == course.id,
                        models.Lesson.order_index == order_index
                    ).first()
                    
                    if lesson:
                        lesson.youtube_id = yt_id
                        count += 1
                        print(f"Updated Lesson {order_index} with YouTube ID {yt_id} ({url})")
                    else:
                        print(f"Lesson {order_index} not found in DB.")
                else:
                    print(f"Row {row}: Could not extract YT ID from {url}")
            else:
                print(f"Row {row}: No URL found in column C.")

        db.commit()
        print(f"Successfully updated {count} lessons with real YouTube IDs.")

    finally:
        db.close()

if __name__ == "__main__":
    import_excel_data()
